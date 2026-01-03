#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
bin_inspector_ui.py
Heuristický analyzátor BIN souborů + UI + export XLSX + konverze.

Závislosti:
  - Python 3.10+
  - (doporučeno) Pillow: pip install pillow
  - openpyxl: pip install openpyxl

Build do EXE (Windows):
  pyinstaller --onefile --noconsole bin_inspector_ui.py
"""

from __future__ import annotations

import os
import sys
import math
import struct
import hashlib
import binascii
import datetime as _dt
from dataclasses import dataclass, field
from typing import Optional, List, Dict, Tuple, Callable, Any

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except Exception as e:
    Workbook = None

try:
    from PIL import Image
except Exception:
    Image = None


# -----------------------------
# Utilities
# -----------------------------

def read_file(path: str, max_bytes: Optional[int] = None) -> bytes:
    with open(path, "rb") as f:
        if max_bytes is None:
            return f.read()
        return f.read(max_bytes)

def sha1_hex(b: bytes) -> str:
    h = hashlib.sha1()
    h.update(b)
    return h.hexdigest()

def entropy_shannon(b: bytes) -> float:
    if not b:
        return 0.0
    freq = [0]*256
    for x in b:
        freq[x] += 1
    n = len(b)
    ent = 0.0
    for c in freq:
        if c:
            p = c / n
            ent -= p * math.log2(p)
    return ent

def is_mostly_printable_ascii(b: bytes, ratio: float = 0.9) -> bool:
    if not b:
        return False
    printable = 0
    for x in b:
        if x in (9, 10, 13) or 32 <= x <= 126:
            printable += 1
    return (printable / len(b)) >= ratio

def u16le(b: bytes, off: int) -> Optional[int]:
    if off+2 > len(b): return None
    return struct.unpack_from("<H", b, off)[0]

def u32le(b: bytes, off: int) -> Optional[int]:
    if off+4 > len(b): return None
    return struct.unpack_from("<I", b, off)[0]

def clamp(x: int, lo: int, hi: int) -> int:
    return lo if x < lo else hi if x > hi else x

def hexdump_prefix(b: bytes, n: int = 64) -> str:
    return binascii.hexlify(b[:n]).decode("ascii")

def safe_makedirs(p: str) -> None:
    os.makedirs(p, exist_ok=True)


# -----------------------------
# Result model
# -----------------------------

@dataclass
class Detection:
    kind: str
    confidence: int
    ext: str = ""
    details: str = ""
    evidence: List[str] = field(default_factory=list)
    convert_hint: str = ""
    raw_image_guess: Optional[Dict[str, Any]] = None  # when kind == RAW_IMAGE_GUESS

@dataclass
class FileReport:
    path: str
    size: int
    sha1: str
    entropy: float
    detections: List[Detection]
    best: Detection


# -----------------------------
# Core detectors (magic bytes etc.)
# -----------------------------

def det_png(b: bytes) -> Optional[Detection]:
    if b.startswith(b"\x89PNG\r\n\x1a\n"):
        return Detection("PNG_IMAGE", 100, ".png", "Valid PNG signature", ["89 50 4E 47 ..."])
    return None

def det_gif(b: bytes) -> Optional[Detection]:
    if b.startswith(b"GIF87a") or b.startswith(b"GIF89a"):
        return Detection("GIF_IMAGE", 100, ".gif", "Valid GIF signature", [b[:6].decode("ascii", "ignore")])
    return None

def det_jpeg(b: bytes) -> Optional[Detection]:
    if len(b) >= 3 and b[0:2] == b"\xFF\xD8" and b[-2:] == b"\xFF\xD9":
        return Detection("JPEG_IMAGE", 95, ".jpg", "SOI/EOI markers present", ["FF D8 ... FF D9"])
    if len(b) >= 2 and b[0:2] == b"\xFF\xD8":
        return Detection("JPEG_IMAGE", 85, ".jpg", "SOI marker present", ["FF D8"])
    return None

def det_bmp(b: bytes) -> Optional[Detection]:
    if b.startswith(b"BM") and len(b) >= 26:
        fsz = u32le(b, 2)
        dib = u32le(b, 14)
        w = u32le(b, 18)
        h = u32le(b, 22)
        ev = ["BM", f"fileSize={fsz}", f"DIB={dib}", f"w={w}", f"h={h}"]
        conf = 90
        if fsz and fsz == len(b):
            conf = 100
        return Detection("BMP_IMAGE", conf, ".bmp", "BMP header", ev)
    return None

def det_ico_cur(b: bytes) -> Optional[Detection]:
    if len(b) >= 6:
        reserved = u16le(b, 0)
        typ = u16le(b, 2)   # 1=ICO, 2=CUR
        count = u16le(b, 4)
        if reserved == 0 and typ in (1, 2) and count and count < 1000:
            kind = "ICON_ICO" if typ == 1 else "CURSOR_CUR"
            ext = ".ico" if typ == 1 else ".cur"
            return Detection(kind, 95, ext, "ICO/CUR header", [f"type={typ}", f"count={count}"])
    return None

def det_wav_riff(b: bytes) -> Optional[Detection]:
    if len(b) >= 12 and b[0:4] == b"RIFF" and b[8:12] == b"WAVE":
        return Detection("WAV_AUDIO", 100, ".wav", "RIFF/WAVE signature", ["RIFF", "WAVE"])
    # other RIFF containers can be extended later
    return None

def det_midi(b: bytes) -> Optional[Detection]:
    if b.startswith(b"MThd") and len(b) >= 14:
        hdr_len = u32le(b, 4)
        fmt = u16le(b, 8)
        ntr = u16le(b, 10)
        div = u16le(b, 12)
        conf = 90
        if hdr_len == 6 and fmt in (0, 1, 2) and ntr and ntr < 256:
            conf = 100
        return Detection("MIDI", conf, ".mid", "MThd header", [f"hdr_len={hdr_len}", f"fmt={fmt}", f"tracks={ntr}", f"div={div}"])
    return None

def det_ogg(b: bytes) -> Optional[Detection]:
    if b.startswith(b"OggS"):
        return Detection("OGG_CONTAINER", 95, ".ogg", "OggS signature", ["OggS"])
    return None

def det_flac(b: bytes) -> Optional[Detection]:
    if b.startswith(b"fLaC"):
        return Detection("FLAC_AUDIO", 100, ".flac", "fLaC signature", ["fLaC"])
    return None

def det_zip(b: bytes) -> Optional[Detection]:
    if b.startswith(b"PK\x03\x04") or b.startswith(b"PK\x05\x06") or b.startswith(b"PK\x07\x08"):
        return Detection("ZIP_ARCHIVE", 95, ".zip", "ZIP signature", ["PK.."])
    return None

def det_id3_mp3(b: bytes) -> Optional[Detection]:
    if b.startswith(b"ID3"):
        return Detection("MP3_AUDIO", 85, ".mp3", "ID3 tag found", ["ID3"])
    # MPEG frame sync heuristic
    if len(b) >= 2:
        if b[0] == 0xFF and (b[1] & 0xE0) == 0xE0:
            return Detection("MP3_AUDIO", 60, ".mp3", "MPEG frame sync heuristic", ["FF Ex"])
    return None

def det_pc_palette(b: bytes) -> Optional[Detection]:
    # Palety co jsme řešili: 96 / 192 / 768 bytes, často v rozsahu 0..63 (VGA 6-bit)
    n = len(b)
    if n in (96, 192, 768):
        mx = max(b) if b else 0
        vga6 = mx <= 63
        # 96 = 32 barev * RGB; 192 = 64 barev; 768 = 256 barev
        colors = n // 3
        conf = 88
        details = f"Palette candidate: {colors} colors (RGB triplets), max={mx}" + (" (VGA6 0..63 likely)" if vga6 else "")
        ev = [f"size={n}", f"colors={colors}", f"max={mx}"]
        if vga6:
            conf = 95
        return Detection("PALETTE_RGB", conf, ".pal", details, ev, convert_hint="Export .pal (raw RGB) and optional .txt dump")
    return None

def det_text(b: bytes) -> Optional[Detection]:
    sample = b[:4096]
    if is_mostly_printable_ascii(sample, 0.93):
        return Detection("TEXT_ASCII", 70, ".txt", "Mostly printable ASCII in first 4KB", ["printable_ratio>=0.93"])
    # UTF-16 LE BOM
    if sample.startswith(b"\xff\xfe"):
        return Detection("TEXT_UTF16LE", 85, ".txt", "UTF-16 LE BOM", ["FF FE"])
    if sample.startswith(b"\xfe\xff"):
        return Detection("TEXT_UTF16BE", 85, ".txt", "UTF-16 BE BOM", ["FE FF"])
    return None

def det_exe_pe(b: bytes) -> Optional[Detection]:
    if b.startswith(b"MZ") and len(b) > 0x40:
        pe_off = u32le(b, 0x3C)
        if pe_off and pe_off+4 <= len(b) and b[pe_off:pe_off+4] == b"PE\0\0":
            return Detection("WIN_PE_EXECUTABLE", 95, ".exe", "MZ + PE header", [f"pe_off=0x{pe_off:X}"])
        return Detection("DOS_MZ", 60, ".bin", "MZ header (no PE found in-range)", ["MZ"])
    return None


# -----------------------------
# RAW image heuristics (Spellcross-like BINs)
# -----------------------------

COMMON_WIDTHS = [
    32, 40, 48, 50, 56, 60, 64, 72, 80, 88, 96, 100, 112, 120, 128, 144, 160,
    176, 192, 200, 208, 224, 240, 256, 272, 288, 300, 320, 336, 352, 360, 384,
    400, 416, 432, 448, 480, 512, 560, 576, 600, 640, 672, 704, 720, 768, 800
]

def _raw_score_8bpp(buf: bytes, w: int, h: int) -> float:
    # Skóre “vypadá jako obrázek”: hladkost hran (lokální korelace)
    # Nižší průměrný rozdíl sousedů => vyšší skóre, ale penalizujeme monotónnost.
    if w <= 1 or h <= 1:
        return -1e9
    # subsample for speed
    step_x = 1 if w <= 320 else 2
    step_y = 1 if h <= 240 else 2

    def at(x: int, y: int) -> int:
        return buf[y*w + x]

    diffs = 0
    cnt = 0
    for y in range(0, h-1, step_y):
        row_off = y*w
        for x in range(0, w-1, step_x):
            a = buf[row_off + x]
            b1 = buf[row_off + x + 1]
            b2 = buf[row_off + x + w]
            diffs += abs(a - b1) + abs(a - b2)
            cnt += 2
    if cnt == 0:
        return -1e9
    avg = diffs / cnt

    # variance proxy: too flat -> penalize
    # quick: sample histogram sparsity
    sample = buf[::max(1, (len(buf)//4096))]
    uniq = len(set(sample))
    flat_pen = 0.0
    if uniq < 16:
        flat_pen = 40.0

    # Entropy proxy (higher entropy is ok but too high suggests compressed/noise)
    ent = entropy_shannon(sample)
    noise_pen = 0.0
    if ent > 7.6:
        noise_pen = (ent - 7.6) * 50.0

    return (255.0 - avg) * 2.0 - flat_pen - noise_pen

def _try_guess_raw_images(b: bytes) -> Optional[Detection]:
    n = len(b)
    if n < 256:
        return None

    # Candidate 8bpp: size must fit w*h
    best = None  # (score, w, h, mode)
    for w in COMMON_WIDTHS:
        if w <= 0:
            continue
        if n % w != 0:
            continue
        h = n // w
        if h <= 0 or h > 4000:
            continue
        # common UI-ish sizes filter
        if (w < 32 or h < 32) and n > 4096:
            continue
        score = _raw_score_8bpp(b, w, h)
        # try swapped orientation (interpret as h*w by transpose preview later)
        if best is None or score > best[0]:
            best = (score, w, h, "8bpp_linear")

    # Candidate 4bpp packed (2 pixels per byte) => pixels = n*2 => w*h == n*2
    best4 = None
    pixels = n * 2
    for w in COMMON_WIDTHS:
        if pixels % w != 0:
            continue
        h = pixels // w
        if h <= 0 or h > 8000:
            continue
        # Build a lightweight expanded buffer for scoring (subsample expand)
        # Expand first min(pixels, 200k) pixels to keep fast
        max_pix = min(pixels, 200_000)
        exp = bytearray(max_pix)
        bi = 0
        pi = 0
        while pi < max_pix and bi < n:
            v = b[bi]
            exp[pi] = (v >> 4) & 0x0F
            if pi+1 < max_pix:
                exp[pi+1] = v & 0x0F
            bi += 1
            pi += 2
        # score on truncated image with w' and h'
        hh = max_pix // w
        if hh < 32:
            continue
        score = _raw_score_8bpp(bytes(exp[:hh*w]), w, hh) - 30.0  # penalize because partial
        if best4 is None or score > best4[0]:
            best4 = (score, w, h, "4bpp_packed")

    # Decide if raw guess is plausible
    chosen = None
    if best is not None and best[0] > 120.0:
        chosen = best
    if best4 is not None and best4[0] > (chosen[0] if chosen else 0) and best4[0] > 90.0:
        chosen = best4

    if not chosen:
        return None

    score, w, h, mode = chosen
    conf = clamp(int(50 + (score / 8.0)), 55, 92)  # heuristic mapping
    details = f"Raw image guess: {mode}, {w}x{h} (bytes={n}), score={score:.1f}"
    ev = [f"w={w}", f"h={h}", f"mode={mode}", f"score={score:.1f}"]

    return Detection(
        "RAW_IMAGE_GUESS",
        conf,
        ".png",
        details,
        ev,
        convert_hint="Convert to PNG (default grayscale or chosen palette)",
        raw_image_guess={"mode": mode, "w": w, "h": h}
    )


# -----------------------------
# Custom detectors placeholder (Spellcross-specific hooks)
# -----------------------------

CUSTOM_DETECTORS: List[Callable[[bytes, str], Optional[Detection]]] = []


def det_spellcross_level_map_raw379x259(b: bytes, path: str) -> Optional[Detection]:
    """
    Spellcross-specific:
    LEVEL_0X.bin appears to be raw indexed 8bpp map chunk stored as full-file
    379x259 = 98161 bytes. No header.
    """
    n = len(b)
    if n != 379 * 259:
        return None
    name = os.path.basename(path).upper()
    # accept LEVEL_02.BIN, LEVEL_0?.BIN etc.
    if not (name.startswith("LEVEL_") and name.endswith(".BIN")):
        # allow also generic .bin of exact size, but lower confidence
        kind = "SPELLCROSS_RAW379x259"
        conf = 88
    else:
        kind = "SPELLCROSS_LEVEL_MAP_379x259"
        conf = 97
    return Detection(
        kind=kind,
        confidence=conf,
        ext=".png",
        details="Spellcross raw indexed image: 379x259, 8bpp (full file, no header).",
        evidence=[f"size={n}", "dims=379x259", "no_header"],
        convert_hint="Convert to PNG using picked palette, or auto <same>.PAL / SYSTEM.PAL"
    )

def det_spellcross_picture_640x480(b: bytes, path: str) -> Optional[Detection]:
    n = len(b)
    if n != 640 * 480:
        return None

    name = os.path.basename(path).upper()
    if name in ("PICTURE.BIN", "INTRO.BIN", "SCREEN.BIN"):
        conf = 98
    else:
        conf = 90

    return Detection(
        kind="SPELLCROSS_PICTURE_640x480",
        confidence=conf,
        ext=".png",
        details="Spellcross fullscreen picture: raw indexed 8bpp 640x480 (no header)",
        evidence=[
            f"size={n}",
            "dims=640x480",
            "indexed_8bpp",
            "fullscreen_ui"
        ],
        convert_hint="Convert to PNG using matching .PAL (same basename) or SYSTEM.PAL"
    )


# ------------------------------------------------------------
# Spellcross ASCII-6bit UI panels (BUY/STATS/INFO/UNITS/OPTIONS/FACTORY/...)
#  - Each byte is usually in 0x40..0x7F and encodes a 6-bit index:
#      idx = byte - 0x40  (0..63)
 # outside that range are treated as 0 (transparent/background)
#  - The file length equals pixel count (1 byte per pixel), so we can factor
#    dimensions exactly within <=640x480.
# ------------------------------------------------------------

def _ascii6_ratio(b: bytes, sample_max: int = 200_000) -> float:
    if not b:
        return 0.0
    s = b[:sample_max]
    hit = 0
    for x in s:
        if 0x40 <= x <= 0x7F:
            hit += 1
    return hit / len(s)

def _ascii6_decode(b: bytes) -> bytes:
    out = bytearray(len(b))
    for i, x in enumerate(b):
        if 0x40 <= x <= 0x7F:
            out[i] = x - 0x40
        else:
            out[i] = 0
    return bytes(out)

def _factor_dims_640x480(n: int) -> List[Tuple[int, int]]:
    dims: List[Tuple[int, int]] = []
    for h in range(1, 481):
        if n % h == 0:
            w = n // h
            if w <= 640:
                dims.append((w, h))
    return dims

def _choose_panel_dims(dims: List[Tuple[int, int]]) -> Tuple[int, int]:
    # Prefer dimensions matching our known UI panel families:
    #  - 406x464 (BUY/UNITS/HIERARCH/RSRCH_BG)
    #  - 569x464 (STATS/OPTIONS/FACTORY)
    #  - 412x464 (INFO)
    def score(w: int, h: int) -> float:
        s = 0.0
        s -= abs(h - 464) * 0.50
        s -= abs(w - 406) * 0.30
        s -= abs(w - 569) * 0.10
        s -= abs(w - 412) * 0.12
        # penalize extreme aspect ratios
        ar = w / h if h else 999.0
        if ar < 0.4 or ar > 2.6:
            s -= 500.0
        return s

    best = None
    for (w, h) in dims:
        if w < 64 or h < 64:
            continue
        sc = score(w, h)
        if best is None or sc > best[0]:
            best = (sc, w, h)
    if best:
        return best[1], best[2]
    return dims[0]

def det_spellcross_ascii6_panel(b: bytes, path: str) -> Optional[Detection]:
    n = len(b)
    if n < 4096:
        return None
    ratio = _ascii6_ratio(b)
    if ratio < 0.60:
        return None
    dims = _factor_dims_640x480(n)
    if not dims:
        return None
    w, h = _choose_panel_dims(dims)
    conf = clamp(int(70 + ratio * 30), 80, 96)
    details = f"Spellcross ASCII-6bit UI panel: {w}x{h} (ascii6_ratio={ratio:.3f})"
    ev = [f"size={n}", f"ascii6_ratio={ratio:.3f}", f"dims={w}x{h}", f"dim_candidates={len(dims)}"]
    return Detection(
        kind="SPELLCROSS_ASCII6_PANEL",
        confidence=conf,
        ext=".png",
        details=details,
        evidence=ev,
        convert_hint="Convert to PNG (ASCII-6bit panel). Uses picked palette / <name>.PAL / SYSTEM.PAL. Exports gray + rotations + transparent.",
        raw_image_guess={"mode": "ascii6_panel", "w": w, "h": h}
    )

def register_custom_detector(fn: Callable[[bytes, str], Optional[Detection]]) -> None:
    CUSTOM_DETECTORS.append(fn)


# Spellcross custom detectors
register_custom_detector(det_spellcross_level_map_raw379x259)
register_custom_detector(det_spellcross_picture_640x480)
register_custom_detector(det_spellcross_ascii6_panel)

# -----------------------------
# Master analysis
# -----------------------------

BASE_DETECTORS: List[Callable[[bytes], Optional[Detection]]] = [
    det_png, det_gif, det_jpeg, det_bmp, det_ico_cur,
    det_wav_riff, det_midi, det_ogg, det_flac, det_zip, det_id3_mp3,
    det_pc_palette, det_exe_pe, det_text
]

def analyze_bytes(b: bytes, path: str) -> FileReport:
    dets: List[Detection] = []
    # Core detectors
    for fn in BASE_DETECTORS:
        try:
            d = fn(b)
            if d:
                dets.append(d)
        except Exception as e:
            dets.append(Detection("DETECTOR_ERROR", 1, "", f"{fn.__name__} failed: {e}", []))

    # Custom detectors
    for fn in CUSTOM_DETECTORS:
        try:
            d = fn(b, path)
            if d:
                dets.append(d)
        except Exception as e:
            dets.append(Detection("CUSTOM_DETECTOR_ERROR", 1, "", f"{fn.__name__} failed: {e}", []))

    # RAW image guess (only if nothing high-confidence already)
    best_conf = max((d.confidence for d in dets), default=0)
    if best_conf < 90:
        try:
            rawd = _try_guess_raw_images(b)
            if rawd:
                dets.append(rawd)
        except Exception:
            pass

    # If still nothing, generic binary classification
    if not dets:
        dets.append(Detection("UNKNOWN_BINARY", 20, ".bin", "No known signatures matched", ["no_magic"]))

    dets.sort(key=lambda d: d.confidence, reverse=True)
    best = dets[0]

    return FileReport(
        path=path,
        size=len(b),
        sha1=sha1_hex(b),
        entropy=entropy_shannon(b[:200_000]),  # limit
        detections=dets,
        best=best
    )


# -----------------------------
# Converters
# -----------------------------

def palette_to_rgb_bytes(pal: bytes) -> bytes:
    # If VGA 6-bit palette, upscale to 8-bit.
    if not pal:
        return pal
    mx = max(pal)
    if mx <= 63:
        out = bytearray(len(pal))
        for i, v in enumerate(pal):
            out[i] = clamp(v * 4, 0, 255)
        return bytes(out)
    return pal

def save_palette_outputs(src_path: str, pal: bytes, out_dir: str) -> List[str]:
    base = os.path.splitext(os.path.basename(src_path))[0]
    safe_makedirs(out_dir)
    out_pal = os.path.join(out_dir, base + ".pal")
    out_txt = os.path.join(out_dir, base + ".pal.txt")
    rgb = palette_to_rgb_bytes(pal)
    with open(out_pal, "wb") as f:
        f.write(rgb)
    # text dump
    cols = len(rgb)//3
    with open(out_txt, "w", encoding="utf-8") as f:
        f.write(f"# {src_path}\n# colors={cols}\n")
        for i in range(cols):
            r, g, b = rgb[i*3:(i+1)*3]
            f.write(f"{i:3d}: {r:3d} {g:3d} {b:3d}\n")
    return [out_pal, out_txt]

def load_palette_file(p: str) -> Optional[bytes]:
    try:
        pal = read_file(p)
        if len(pal) in (96, 192, 768):
            return palette_to_rgb_bytes(pal)
        # allow text? no, keep simple
        return None
    except Exception:
        return None


def _expand_palette_to_256(pal_rgb: bytes) -> Optional[bytes]:
    """
    Accepts palette bytes in sizes:
      - 96  (32 colors * RGB)
      - 192 (64 colors * RGB)
      - 768 (256 colors * RGB)
    Returns a 768-byte palette suitable for PIL putpalette().
    Strategy:
      - 768 -> return as-is
      - 192 -> repeat 64-color entries to fill 256
      - 96  -> repeat 32-color entries to fill 256
    """
    if not pal_rgb:
        return None
    if len(pal_rgb) == 768:
        return pal_rgb[:768]
    if len(pal_rgb) == 192:
        cols = [pal_rgb[i*3:(i+1)*3] for i in range(64)]
        out = bytearray()
        for i in range(256):
            out += cols[i % 64]
        return bytes(out[:768])
    if len(pal_rgb) == 96:
        cols = [pal_rgb[i*3:(i+1)*3] for i in range(32)]
        out = bytearray()
        for i in range(256):
            out += cols[i % 32]
        return bytes(out[:768])
    return None

def _try_auto_palette_for(path: str, ui_palette_path: Optional[str]) -> Optional[bytes]:
    """
    Palette selection priority:
      1) palette picked in UI
      2) sibling palette with same basename: <name>.PAL / <name>.pal
      3) fallback: SYSTEM.PAL in same folder
    Returns palette in original size (96/192/768), not expanded.
    """
    if ui_palette_path:
        pal = load_palette_file(ui_palette_path)
        if pal:
            return pal
    base = os.path.splitext(path)[0]
    for ext in (".PAL", ".pal"):
        p = base + ext
        if os.path.isfile(p):
            pal = load_palette_file(p)
            if pal:
                return pal
    sys_pal = os.path.join(os.path.dirname(path), "SYSTEM.PAL")
    if os.path.isfile(sys_pal):
        pal = load_palette_file(sys_pal)
        if pal:
            return pal
    return None

def _save_raw_indexed_png(src_path: str, pixels: bytes, w: int, h: int, out_dir: str, pal_rgb: Optional[bytes]) -> Optional[str]:
    """
    Save raw 8bpp indexed image to PNG using provided palette.
    If palette is 32/64 colors, it will be expanded to 256 by repeating.
    """
    if Image is None:
        return None
    safe_makedirs(out_dir)
    base = os.path.splitext(os.path.basename(src_path))[0]
    out_png = os.path.join(out_dir, base + f"_{w}x{h}_raw.png")

    img = Image.frombytes("L", (w, h), pixels[:w*h])
    pal256 = _expand_palette_to_256(pal_rgb) if pal_rgb else None
    if pal256:
        pimg = img.convert("P")
        pimg.putpalette(pal256)
        pimg.save(out_png)
    else:
        # grayscale debug
        img.save(out_png)
    return out_png


def raw_image_to_png(src_path: str, b: bytes, guess: Dict[str, Any], out_dir: str, palette_rgb: Optional[bytes]) -> Optional[str]:
    if Image is None:
        return None
    w = int(guess["w"])
    h = int(guess["h"])
    mode = guess["mode"]
    safe_makedirs(out_dir)
    base = os.path.splitext(os.path.basename(src_path))[0]
    out_png = os.path.join(out_dir, base + f"_{w}x{h}_{mode}.png")

    if mode == "8bpp_linear":
        img = Image.frombytes("L", (w, h), b[:w*h])
        if palette_rgb and len(palette_rgb) == 768:
            # paletted
            pimg = img.convert("P")
            pimg.putpalette(palette_rgb)
            pimg.save(out_png)
        else:
            img.save(out_png)
        return out_png

    if mode == "4bpp_packed":
        # Expand to 8-bit indices 0..15
        pixels = w*h
        needed_bytes = (pixels + 1)//2
        packed = b[:needed_bytes]
        exp = bytearray(pixels)
        pi = 0
        for v in packed:
            if pi < pixels:
                exp[pi] = (v >> 4) & 0x0F
                pi += 1
            if pi < pixels:
                exp[pi] = v & 0x0F
                pi += 1
        img = Image.frombytes("L", (w, h), bytes(exp))
        # If user gave 256-color palette, we can still map first 16 entries. If 16/64 palettes are given, OK.
        if palette_rgb:
            # Build a 256 palette even if smaller
            pal = palette_rgb
            if len(pal) == 48:   # 16 colors
                pal256 = pal + bytes([0, 0, 0]) * (256 - 16)
            elif len(pal) == 192:  # 64 colors
                pal256 = pal + bytes([0, 0, 0]) * (256 - 64)
            elif len(pal) == 768:
                pal256 = pal
            else:
                pal256 = None
            if pal256:
                pimg = img.convert("P")
                pimg.putpalette(pal256)
                pimg.save(out_png)
                return out_png
        img.save(out_png)
        return out_png

    return None


def _save_spellcross_ascii6_panel_png(src_path: str, b: bytes, w: int, h: int, out_dir: str, pal_rgb: Optional[bytes]) -> List[str]:
    """
    Decode Spellcross ASCII-6bit panel (0x40..0x7F -> 0..63) and export:
      - base PNG (paletted if palette available, else grayscale)
      - gray PNG (debug)
      - rot90 / rot-90 / rot180 PNG
      - transparent PNG (index 0 as alpha=0)

    Returns list of written paths.
    """
    if Image is None:
        return []

    safe_makedirs(out_dir)
    base = os.path.splitext(os.path.basename(src_path))[0]

    idx = _ascii6_decode(b)[:w*h]

    pal256 = _expand_palette_to_256(pal_rgb) if pal_rgb else None

    outs: List[str] = []

    # gray (always)
    gray = bytes((min(255, v * 4) for v in idx))
    out_gray = os.path.join(out_dir, base + f"_{w}x{h}_ascii6_gray.png")
    Image.frombytes("L", (w, h), gray).save(out_gray)
    outs.append(out_gray)

    # base
    out_png = os.path.join(out_dir, base + f"_{w}x{h}_ascii6.png")
    if pal256:
        im_base = Image.frombytes("P", (w, h), idx)
        im_base.putpalette(pal256)
        im_base.save(out_png)
    else:
        im_base = Image.frombytes("L", (w, h), gray)
        im_base.save(out_png)
    outs.append(out_png)

    # rotations
    for ang in (90, -90, 180):
        out_r = os.path.join(out_dir, base + f"_{w}x{h}_ascii6_rot{ang}.png")
        im_base.rotate(ang, expand=True).save(out_r)
        outs.append(out_r)

    # transparent (only meaningful with palette)
    if pal256:
        out_t = os.path.join(out_dir, base + f"_{w}x{h}_ascii6_transparent.png")
        imP = Image.frombytes("P", (w, h), idx)
        imP.putpalette(pal256)
        rgba = imP.convert("RGBA")
        alpha = bytearray(w * h)
        for i, v in enumerate(idx):
            alpha[i] = 0 if v == 0 else 255
        rgba.putalpha(Image.frombytes("L", (w, h), bytes(alpha)))
        rgba.save(out_t)
        outs.append(out_t)

    return outs

def unwrap_known(src_path: str, out_dir: str, ext: str) -> str:
    safe_makedirs(out_dir)
    base = os.path.splitext(os.path.basename(src_path))[0]
    out_path = os.path.join(out_dir, base + ext)
    # copy bytes
    with open(src_path, "rb") as fi, open(out_path, "wb") as fo:
        fo.write(fi.read())
    return out_path


# -----------------------------
# XLSX Export
# -----------------------------

def export_xlsx(reports: List[FileReport], out_path: str) -> None:
    if Workbook is None:
        raise RuntimeError("openpyxl is not installed")
    wb = Workbook()
    ws = wb.active
    ws.title = "BIN Report"

    headers = [
        "name", "path", "size", "sha1", "entropy",
        "best_kind", "best_conf", "best_ext", "best_details",
        "evidence", "all_candidates"
    ]
    ws.append(headers)

    for r in reports:
        best = r.best
        evidence = "; ".join(best.evidence[:12])
        allc = " | ".join([f"{d.kind}:{d.confidence}{d.ext}" for d in r.detections[:8]])
        ws.append([
            os.path.basename(r.path), r.path, r.size, r.sha1, round(r.entropy, 3),
            best.kind, best.confidence, best.ext, best.details,
            evidence, allc
        ])

    # autosize
    for col in range(1, len(headers)+1):
        maxlen = 10
        for row in range(1, ws.max_row+1):
            val = ws.cell(row=row, column=col).value
            if val is None:
                continue
            maxlen = max(maxlen, len(str(val)))
        ws.column_dimensions[get_column_letter(col)].width = min(60, maxlen + 2)

    wb.save(out_path)


# -----------------------------
# UI
# -----------------------------

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("BIN Inspector (heuristics + XLSX + convert)")
        self.geometry("1180x720")
        self.minsize(980, 620)

        self.files: List[str] = []
        self.reports: Dict[str, FileReport] = {}
        self.palette_path: Optional[str] = None
        self.output_dir: str = os.path.abspath(os.path.join(os.getcwd(), "bin_out"))

        self._build_ui()

    def _build_ui(self):
        top = ttk.Frame(self)
        top.pack(side=tk.TOP, fill=tk.X, padx=10, pady=10)

        ttk.Button(top, text="Add files…", command=self.add_files).pack(side=tk.LEFT)
        ttk.Button(top, text="Add folder…", command=self.add_folder).pack(side=tk.LEFT, padx=(8, 0))
        ttk.Button(top, text="Clear", command=self.clear_all).pack(side=tk.LEFT, padx=(8, 0))

        ttk.Separator(top, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=12)

        ttk.Button(top, text="Analyze", command=self.analyze_all).pack(side=tk.LEFT)
        ttk.Button(top, text="Export XLSX…", command=self.export_xlsx_ui).pack(side=tk.LEFT, padx=(8, 0))
        ttk.Button(top, text="Convert selected", command=self.convert_selected).pack(side=tk.LEFT, padx=(8, 0))
        ttk.Button(top, text="Convert all", command=self.convert_all).pack(side=tk.LEFT, padx=(8, 0))

        ttk.Separator(top, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=12)

        ttk.Label(top, text="Output:").pack(side=tk.LEFT)
        self.out_var = tk.StringVar(value=self.output_dir)
        out_entry = ttk.Entry(top, textvariable=self.out_var, width=45)
        out_entry.pack(side=tk.LEFT, padx=(6, 0))
        ttk.Button(top, text="…", width=3, command=self.pick_output).pack(side=tk.LEFT, padx=(4, 0))

        ttk.Separator(top, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=12)

        ttk.Label(top, text="Palette (optional):").pack(side=tk.LEFT)
        self.pal_var = tk.StringVar(value="")
        pal_entry = ttk.Entry(top, textvariable=self.pal_var, width=28)
        pal_entry.pack(side=tk.LEFT, padx=(6, 0))
        ttk.Button(top, text="Pick", command=self.pick_palette).pack(side=tk.LEFT, padx=(4, 0))
        ttk.Button(top, text="Clear", command=self.clear_palette).pack(side=tk.LEFT, padx=(4, 0))

        mid = ttk.PanedWindow(self, orient=tk.HORIZONTAL)
        mid.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))

        left = ttk.Frame(mid)
        right = ttk.Frame(mid)
        mid.add(left, weight=2)
        mid.add(right, weight=1)

        # Tree view
        cols = ("name", "type", "conf", "size", "entropy", "ext")
        self.tree = ttk.Treeview(left, columns=cols, show="headings", selectmode="extended")
        self.tree.heading("name", text="File")
        self.tree.heading("type", text="Best type")
        self.tree.heading("conf", text="Conf")
        self.tree.heading("size", text="Bytes")
        self.tree.heading("entropy", text="Entropy")
        self.tree.heading("ext", text="Ext")

        self.tree.column("name", width=220, anchor="w")
        self.tree.column("type", width=280, anchor="w")
        self.tree.column("conf", width=60, anchor="center")
        self.tree.column("size", width=90, anchor="e")
        self.tree.column("entropy", width=90, anchor="e")
        self.tree.column("ext", width=60, anchor="center")

        vsb = ttk.Scrollbar(left, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)

        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)

        self.tree.bind("<<TreeviewSelect>>", self.on_select)

        # Details panel
        self.detail = tk.Text(right, wrap="word", height=20)
        self.detail.pack(fill=tk.BOTH, expand=True)

        bot = ttk.Frame(self)
        bot.pack(side=tk.BOTTOM, fill=tk.X, padx=10, pady=8)

        self.status_var = tk.StringVar(value="Ready.")
        ttk.Label(bot, textvariable=self.status_var).pack(side=tk.LEFT)

    def set_status(self, s: str):
        self.status_var.set(s)
        self.update_idletasks()

    def add_files(self):
        paths = filedialog.askopenfilenames(title="Select files", filetypes=[("All files", "*.*")])
        if not paths:
            return
        for p in paths:
            if p not in self.files:
                self.files.append(p)
        self.refresh_tree()

    def add_folder(self):
        folder = filedialog.askdirectory(title="Select folder")
        if not folder:
            return
        # Add common bin-like files first, but allow all
        added = 0
        for root, _, files in os.walk(folder):
            for fn in files:
                p = os.path.join(root, fn)
                if p not in self.files:
                    self.files.append(p)
                    added += 1
        self.refresh_tree()
        self.set_status(f"Added {added} files from folder.")

    def clear_all(self):
        self.files.clear()
        self.reports.clear()
        for i in self.tree.get_children():
            self.tree.delete(i)
        self.detail.delete("1.0", tk.END)
        self.set_status("Cleared.")

    def refresh_tree(self):
        # keep existing analysis if possible
        for i in self.tree.get_children():
            self.tree.delete(i)
        for p in self.files:
            rep = self.reports.get(p)
            if rep:
                best = rep.best
                self.tree.insert("", tk.END, iid=p, values=(os.path.basename(p), best.kind, best.confidence, rep.size, f"{rep.entropy:.3f}", best.ext))
            else:
                self.tree.insert("", tk.END, iid=p, values=(os.path.basename(p), "—", "—", os.path.getsize(p), "—", ""))
        self.set_status(f"Files: {len(self.files)}")

    def analyze_all(self):
        if not self.files:
            messagebox.showinfo("BIN Inspector", "No files selected.")
            return
        self.set_status("Analyzing…")
        ok = 0
        for idx, p in enumerate(self.files, 1):
            try:
                b = read_file(p)
                rep = analyze_bytes(b, p)
                self.reports[p] = rep
                ok += 1
            except Exception as e:
                self.reports[p] = FileReport(
                    path=p, size=os.path.getsize(p), sha1="",
                    entropy=0.0,
                    detections=[Detection("READ_ERROR", 1, "", str(e), [])],
                    best=Detection("READ_ERROR", 1, "", str(e), [])
                )
            if idx % 10 == 0:
                self.set_status(f"Analyzing… {idx}/{len(self.files)}")
        self.refresh_tree()
        self.set_status(f"Analyzed: {ok}/{len(self.files)}")

    def on_select(self, _evt=None):
        sel = self.tree.selection()
        self.detail.delete("1.0", tk.END)
        if not sel:
            return
        p = sel[0]
        rep = self.reports.get(p)
        self.detail.insert(tk.END, f"Path: {p}\n")
        self.detail.insert(tk.END, f"Size: {os.path.getsize(p)} bytes\n")
        if rep:
            self.detail.insert(tk.END, f"SHA1: {rep.sha1}\n")
            self.detail.insert(tk.END, f"Entropy: {rep.entropy:.3f}\n\n")
            self.detail.insert(tk.END, f"BEST: {rep.best.kind} ({rep.best.confidence}%) {rep.best.ext}\n")
            self.detail.insert(tk.END, f"Details: {rep.best.details}\n")
            if rep.best.convert_hint:
                self.detail.insert(tk.END, f"Convert: {rep.best.convert_hint}\n")
            if rep.best.evidence:
                self.detail.insert(tk.END, "Evidence:\n")
                for ev in rep.best.evidence:
                    self.detail.insert(tk.END, f"  - {ev}\n")

            self.detail.insert(tk.END, "\nCandidates:\n")
            for d in rep.detections[:12]:
                self.detail.insert(tk.END, f"  {d.kind:18s} {d.confidence:3d}% {d.ext:5s}  {d.details}\n")
        else:
            self.detail.insert(tk.END, "\n(Not analyzed yet.)\n")

    def pick_output(self):
        d = filedialog.askdirectory(title="Pick output directory")
        if not d:
            return
        self.output_dir = d
        self.out_var.set(d)

    def pick_palette(self):
        p = filedialog.askopenfilename(title="Select palette file", filetypes=[("Palette / any", "*.*")])
        if not p:
            return
        self.palette_path = p
        self.pal_var.set(p)

    def clear_palette(self):
        self.palette_path = None
        self.pal_var.set("")

    def export_xlsx_ui(self):
        if not self.reports:
            messagebox.showinfo("BIN Inspector", "Nothing analyzed yet.")
            return
        if Workbook is None:
            messagebox.showerror("Missing dependency", "openpyxl is not installed. Install: pip install openpyxl")
            return
        out = filedialog.asksaveasfilename(
            title="Save XLSX report",
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx")]
        )
        if not out:
            return
        try:
            export_xlsx(list(self.reports.values()), out)
            messagebox.showinfo("Export", f"Saved:\n{out}")
        except Exception as e:
            messagebox.showerror("Export failed", str(e))

    def _convert_one(self, p: str) -> Tuple[bool, str]:
        rep = self.reports.get(p)
        if not rep:
            return False, "Not analyzed"
        best = rep.best
        out_dir = self.out_var.get().strip() or self.output_dir

        # Known formats: unwrap/copy with ext
        if best.kind in ("PNG_IMAGE", "GIF_IMAGE", "JPEG_IMAGE", "BMP_IMAGE", "ICON_ICO", "CURSOR_CUR",
                         "WAV_AUDIO", "MIDI", "OGG_CONTAINER", "FLAC_AUDIO", "ZIP_ARCHIVE", "MP3_AUDIO",
                         "WIN_PE_EXECUTABLE", "TEXT_ASCII", "TEXT_UTF16LE", "TEXT_UTF16BE"):
            outp = unwrap_known(p, out_dir, best.ext or ".bin")
            return True, f"Saved {outp}"

        # Palette
        if best.kind == "PALETTE_RGB":
            b = read_file(p)
            outs = save_palette_outputs(p, b, out_dir)
            return True, "Saved " + ", ".join(outs)


        # Spellcross: ASCII-6bit UI panel (BUY/STATS/INFO/UNITS/OPTIONS/FACTORY/...)
        if best.kind == "SPELLCROSS_ASCII6_PANEL":
            if Image is None:
                return False, "Pillow not installed (pip install pillow)"
            b = read_file(p)
            pal_raw = _try_auto_palette_for(p, self.palette_path)
            guess = best.raw_image_guess or {}
            w = int(guess.get("w", 0) or 0)
            h = int(guess.get("h", 0) or 0)
            if w <= 0 or h <= 0 or w * h > len(b):
                dims = _factor_dims_640x480(len(b))
                if not dims:
                    return False, "Could not determine panel dimensions"
                w, h = _choose_panel_dims(dims)
            outs = _save_spellcross_ascii6_panel_png(p, b, w, h, out_dir, pal_raw)
            if not outs:
                return False, "ASCII-6bit panel convert failed"
            return True, "Saved " + ", ".join(outs)

        # Raw image guess
        if best.kind == "RAW_IMAGE_GUESS":
            if Image is None:
                return False, "Pillow not installed (pip install pillow)"
            pal_rgb = None
            if self.palette_path:
                pal_rgb = load_palette_file(self.palette_path)
            b = read_file(p)
            outp = raw_image_to_png(p, b, best.raw_image_guess or {}, out_dir, pal_rgb)
            if not outp:
                return False, "Raw image convert failed"
            return True, f"Saved {outp}"


        # Spellcross: raw indexed map 379x259 full-file
        if best.kind in ("SPELLCROSS_LEVEL_MAP_379x259", "SPELLCROSS_RAW379x259"):
            w, h = 379, 259
            pixels = read_file(p)
            pal_raw = _try_auto_palette_for(p, self.palette_path)
            outp = _save_raw_indexed_png(p, pixels, w, h, out_dir, pal_raw)
            if not outp:
                return False, "Pillow not installed (pip install pillow)"
            return True, f"Saved {outp}"

        if best.kind == "SPELLCROSS_PICTURE_640x480":
            w, h = 640, 480
            pixels = read_file(p)
            pal_raw = _try_auto_palette_for(p, self.palette_path)
            outp = _save_raw_indexed_png(p, pixels, w, h, out_dir, pal_raw)
            return True, f"Saved {outp}"

        return False, "No converter for this type"

    def convert_selected(self):
        sel = list(self.tree.selection())
        if not sel:
            messagebox.showinfo("Convert", "No selection.")
            return
        if not self.reports:
            messagebox.showinfo("Convert", "Analyze first.")
            return
        ok = 0
        failed = 0
        msgs = []
        for p in sel:
            success, msg = self._convert_one(p)
            if success:
                ok += 1
            else:
                failed += 1
            msgs.append(f"{os.path.basename(p)}: {msg}")
        messagebox.showinfo("Convert result", f"OK: {ok}, Failed: {failed}\n\n" + "\n".join(msgs[:30]))

    def convert_all(self):
        if not self.files:
            messagebox.showinfo("Convert", "No files.")
            return
        if not self.reports:
            messagebox.showinfo("Convert", "Analyze first.")
            return
        ok = 0
        failed = 0
        for idx, p in enumerate(self.files, 1):
            success, _ = self._convert_one(p)
            ok += 1 if success else 0
            failed += 0 if success else 1
            if idx % 10 == 0:
                self.set_status(f"Converting… {idx}/{len(self.files)} (OK={ok}, Fail={failed})")
        self.set_status(f"Convert done. OK={ok}, Fail={failed}")
        messagebox.showinfo("Convert", f"Done.\nOK: {ok}\nFailed: {failed}")

def main():
    # CLI fallback: python bin_inspector_ui.py file1 file2 ...  (prints report)
    if len(sys.argv) > 1 and sys.argv[1] != "--ui":
        paths = [p for p in sys.argv[1:] if os.path.isfile(p)]
        if not paths:
            print("No valid files.")
            return 2
        for p in paths:
            b = read_file(p)
            rep = analyze_bytes(b, p)
            best = rep.best
            print(f"\n{p}")
            print(f"  size={rep.size} sha1={rep.sha1} entropy={rep.entropy:.3f}")
            print(f"  BEST: {best.kind} {best.confidence}% {best.ext} :: {best.details}")
            for d in rep.detections[:8]:
                print(f"    - {d.kind:18s} {d.confidence:3d}% {d.ext:5s} {d.details}")
        return 0

    app = App()
    app.mainloop()
    return 0

if __name__ == "__main__":
    raise SystemExit(main())
